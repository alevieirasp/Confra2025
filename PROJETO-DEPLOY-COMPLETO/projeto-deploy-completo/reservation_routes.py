from flask import Blueprint, request, jsonify
from src.models import db
from src.models.reservation import Reservation
import csv
import io
from flask import make_response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import tempfile
import os

reservation_bp = Blueprint("reservation", __name__)

@reservation_bp.route("/reservations", methods=["GET"])
def get_reservations():
    """Retorna todas as reservas"""
    try:
        reservations = Reservation.query.all()
        return jsonify([reservation.to_dict() for reservation in reservations])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@reservation_bp.route("/reservations", methods=["POST"])
def create_reservation():
    """Cria uma nova reserva"""
    try:
        data = request.get_json()
        
        # Validação dos dados
        if not data or not all(key in data for key in ["bike_id", "name", "shirt_size", "shirt_name"]):
            return jsonify({"error": "Dados incompletos"}), 400
        
        # Verifica se a bike já está reservada
        existing_reservation = Reservation.query.filter_by(bike_id=data["bike_id"]).first()
        if existing_reservation:
            return jsonify({"error": "Bike já está reservada"}), 409
        
        # Cria a nova reserva
        reservation = Reservation(
            bike_id=data["bike_id"],
            name=data["name"],
            shirt_size=data["shirt_size"],
            shirt_name=data["shirt_name"]
        )
        
        db.session.add(reservation)
        db.session.commit()
        
        return jsonify(reservation.to_dict()), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@reservation_bp.route("/reservations/<int:bike_id>", methods=["DELETE"])
def delete_reservation(bike_id):
    """Remove uma reserva"""
    try:
        # Debug: verificar se a reserva existe
        reservation = Reservation.query.filter_by(bike_id=bike_id).first()
        if not reservation:
            # Listar todas as reservas para debug
            all_reservations = Reservation.query.all()
            bike_ids = [r.bike_id for r in all_reservations]
            return jsonify({
                "error": f"Reserva não encontrada para bike {bike_id}",
                "existing_bikes": bike_ids,
                "total_reservations": len(all_reservations)
            }), 404
        
        db.session.delete(reservation)
        db.session.commit()
        
        return jsonify({"message": "Reserva removida com sucesso"}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@reservation_bp.route("/reservations/export", methods=["GET"])
def export_reservations():
    """Exporta as reservas em formato CSV"""
    try:
        reservations = Reservation.query.all()
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Cabeçalho
        writer.writerow(["Bike", "Nome", "Tamanho da Camiseta", "Nome na Camiseta", "Data/Hora da Reserva"])
        
        # Dados
        for reservation in reservations:
            writer.writerow([
                f"Bike {reservation.bike_id}",
                reservation.name,
                reservation.shirt_size,
                reservation.shirt_name,
                reservation.timestamp.strftime("%d/%m/%Y, %H:%M:%S")
            ])
        
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers["Content-Type"] = "text/csv"
        response.headers["Content-Disposition"] = "attachment; filename=reservas-bikes.csv"
        
        return response
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@reservation_bp.route("/reservations/export-excel", methods=["GET"])
def export_reservations_excel():
    """Exporta as reservas em formato Excel (.xlsx)"""
    try:
        reservations = Reservation.query.all()
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reservas Bikes - Aula Especial"
        
        # Cabeçalhos
        headers = ["Bike", "Nome Completo", "Tamanho da Camiseta", "Nome na Camiseta", "Data da Reserva", "Hora da Reserva"]
        ws.append(headers)
        
        # Estilizar cabeçalhos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Adicionar dados
        for reservation in reservations:
            ws.append([
                f"Bike {reservation.bike_id:02d}",
                reservation.name,
                reservation.shirt_size,
                reservation.shirt_name,
                reservation.timestamp.strftime("%d/%m/%Y"),
                reservation.timestamp.strftime("%H:%M:%S")
            ])
        
        # Ajustar largura das colunas
        column_widths = [12, 25, 20, 25, 15, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Adicionar bordas e alinhamento aos dados
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(reservations) + 1):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Não aplicar aos cabeçalhos
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Adicionar informações adicionais
        ws.append([])  # Linha vazia
        ws.append(["Total de Reservas:", len(reservations)])
        ws.append(["Data de Exportação:", reservation.timestamp.strftime("%d/%m/%Y %H:%M:%S") if reservations else "N/A"])
        ws.append(["Sistema:", "Reservas Aula Especial #Veloucos"])
        
        # Salvar em arquivo temporário
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Ler o arquivo e retornar como resposta
        with open(temp_file.name, 'rb') as f:
            excel_data = f.read()
        
        # Limpar arquivo temporário
        os.unlink(temp_file.name)
        
        response = make_response(excel_data)
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        response.headers["Content-Disposition"] = f"attachment; filename=reservas-bikes-veloucos.xlsx"
        
        return response
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@reservation_bp.route("/reservations/clear", methods=["DELETE"])
def clear_reservations():
    """Limpa todas as reservas"""
    try:
        num_rows_deleted = db.session.query(Reservation).delete()
        db.session.commit()
        return jsonify({"message": f"{num_rows_deleted} reservas removidas com sucesso"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@reservation_bp.route("/reservations/debug", methods=["GET"])
def debug_reservations():
    """Endpoint para debug - lista todas as reservas com detalhes"""
    try:
        reservations = Reservation.query.all()
        debug_info = []
        for reservation in reservations:
            debug_info.append({
                "id": reservation.id,
                "bike_id": reservation.bike_id,
                "name": reservation.name,
                "shirt_size": reservation.shirt_size,
                "timestamp": reservation.timestamp.isoformat()
            })
        return jsonify({
            "total_reservations": len(reservations),
            "reservations": debug_info
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

